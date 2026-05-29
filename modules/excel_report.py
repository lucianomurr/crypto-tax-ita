"""
Genera il file Excel nel formato standard (Riassunto + foglio per exchange).
Struttura: foglio "Riassunto" + foglio per exchange con dettaglio asset.
"""
import io
from datetime import date, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Palette colori ────────────────────────────────────────────────────────────
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # blu scuro
_ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")   # azzurro chiaro
_TOTAL_FILL   = PatternFill("solid", fgColor="2E75B6")   # blu medio
_WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_TOTAL_FONT   = Font(bold=True, color="FFFFFF", size=10)
_BODY_FONT    = Font(size=10)
_THIN         = Side(style="thin", color="BFBFBF")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT        = Alignment(horizontal="right",  vertical="center")
_LEFT         = Alignment(horizontal="left",   vertical="center")


def _giorni_detenzione(asset: str, df: pd.DataFrame, tax_year: int,
                        qty_start: float, qty_end: float) -> int:
    """Calcola i giorni di effettiva detenzione nell'anno."""
    start_of_year = date(tax_year, 1, 1)
    end_of_year   = date(tax_year, 12, 31)

    asset_df = df[df["asset"] == asset]
    buy_types  = {"BUY", "RECEIVE", "STAKING", "EARN"}
    sell_types = {"SELL", "SEND", "CONVERT"}

    if qty_start > 0:
        hold_start = start_of_year
    else:
        # Prima acquisizione nell'anno
        first_buy = asset_df[
            asset_df["tx_type"].isin(buy_types) &
            asset_df["date"].apply(lambda d: d.year == tax_year)
        ]["date"].min()
        hold_start = first_buy if not pd.isna(first_buy) else start_of_year

    if qty_end > 0:
        hold_end = end_of_year
    else:
        # Ultima vendita nell'anno
        last_sell = asset_df[
            asset_df["tx_type"].isin(sell_types) &
            asset_df["date"].apply(lambda d: d.year == tax_year)
        ]["date"].max()
        hold_end = last_sell if not pd.isna(last_sell) else end_of_year

    return max(1, (hold_end - hold_start).days + 1)


def _style_header(ws, row_idx: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.border    = _BORDER
        cell.alignment = _CENTER


def _style_row(ws, row_idx: int, n_cols: int, is_alt: bool, is_total: bool) -> None:
    fill = _TOTAL_FILL if is_total else (_ALT_FILL if is_alt else _WHITE_FILL)
    font = _TOTAL_FONT if is_total else _BODY_FONT
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill   = fill
        cell.font   = font
        cell.border = _BORDER


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 30)


def generate_rw_excel(
    rw_df: pd.DataFrame,
    df_transactions: pd.DataFrame,
    tax_year: int,
    exchange_label: str = "Coinbase",
) -> io.BytesIO:
    """
    Genera il file Excel RW nel formato standard.

    Parametri:
        rw_df            – output di compute_rw_data()
        df_transactions  – DataFrame transazioni originale (per calcolo giorni)
        tax_year         – anno d'imposta
        exchange_label   – nome exchange (es. "Coinbase")

    Ritorna un BytesIO pronto per st.download_button.
    """
    start_label = f"01-01-{tax_year}"
    end_label   = f"31-12-{tax_year}"

    # ── Costruzione righe dettaglio ───────────────────────────────────────────
    detail_rows = []
    for _, row in rw_df.iterrows():
        asset      = row["Asset"]
        qty_start  = float(row["Quantita_01gen"])
        qty_end    = float(row["Quantita_31dic"])
        val_start  = float(row["Valore_Iniziale_EUR"])
        val_end    = float(row["Valore_Finale_EUR"])
        price_start = float(row.get("Prezzo_01gen_EUR", 0) or 0)
        price_end   = float(row.get("Prezzo_31dic_EUR", 0) or 0)

        giorni    = _giorni_detenzione(asset, df_transactions, tax_year, qty_start, qty_end)
        ivaca_02  = round(val_end * 0.002, 2)
        ic_giorni = round(ivaca_02 * giorni / 365, 2)

        detail_rows.append({
            "Etichetta":                          exchange_label,
            "Simbolo":                            asset,
            "Data Inizio detenzione":             start_label,
            "Quantità (inizio)":                  qty_start,
            "Prezzo alla data (inizio)":          price_start,
            "Valore inizio detenzione":           round(val_start, 2),
            "Data ultimo giorno detenzione":      end_label,
            "Quantità (fine)":                    qty_end,
            "Prezzo alla data (fine)":            price_end,
            "Valore fine detenzione":             round(val_end, 2),
            "0,2% Del Valore Finale RW":          ivaca_02,
            "Giorni detenzione":                  giorni,
            "IC rapportato ai giorni detenzione": ic_giorni,
        })

    # Riga Complessivo nel foglio dettaglio
    tot_val_start = sum(r["Valore inizio detenzione"] for r in detail_rows)
    tot_val_end   = sum(r["Valore fine detenzione"] for r in detail_rows)
    tot_ivaca     = round(tot_val_end * 0.002, 2)
    tot_ic        = sum(r["IC rapportato ai giorni detenzione"] for r in detail_rows)
    detail_rows.append({
        "Etichetta":                          "Complessivo",
        "Simbolo":                            None,
        "Data Inizio detenzione":             None,
        "Quantità (inizio)":                  None,
        "Prezzo alla data (inizio)":          None,
        "Valore inizio detenzione":           round(tot_val_start, 2),
        "Data ultimo giorno detenzione":      None,
        "Quantità (fine)":                    None,
        "Prezzo alla data (fine)":            None,
        "Valore fine detenzione":             round(tot_val_end, 2),
        "0,2% Del Valore Finale RW":          tot_ivaca,
        "Giorni detenzione":                  None,
        "IC rapportato ai giorni detenzione": round(tot_ic, 2),
    })

    # ── Costruzione righe Riassunto ───────────────────────────────────────────
    summary_rows = [
        {
            "Etichetta":                          exchange_label,
            "Controvalore inizio detenzione":     round(tot_val_start, 2),
            "Data Inizio detenzione":             start_label,
            "Valore Iniziale RW":                 round(tot_val_start),
            "Controvalore fine detenzione":       round(tot_val_end, 2),
            "Data ultimo giorno detenzione":      end_label,
            "Valore Finale RW":                   round(tot_val_end),
            "0,2% Del Valore Finale RW":          tot_ivaca,
            "Giorni detenzione":                  365,
            "IC rapportato ai giorni detenzione": round(tot_ic, 2),
        },
        {
            "Etichetta":                          "Complessivo",
            "Controvalore inizio detenzione":     round(tot_val_start, 2),
            "Data Inizio detenzione":             None,
            "Valore Iniziale RW":                 round(tot_val_start),
            "Controvalore fine detenzione":       round(tot_val_end, 2),
            "Data ultimo giorno detenzione":      None,
            "Valore Finale RW":                   round(tot_val_end),
            "0,2% Del Valore Finale RW":          tot_ivaca,
            "Giorni detenzione":                  None,
            "IC rapportato ai giorni detenzione": round(tot_ic, 2),
        },
    ]

    # ── Scrittura Excel ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Riassunto",       index=False)
        pd.DataFrame(detail_rows).to_excel(writer,  sheet_name=exchange_label[:31], index=False)

    # ── Styling ───────────────────────────────────────────────────────────────
    buf.seek(0)
    wb = load_workbook(buf)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        n_cols = ws.max_column

        # Header row
        _style_header(ws, 1, n_cols)
        ws.row_dimensions[1].height = 36

        # Data rows
        for r_idx in range(2, ws.max_row + 1):
            first_cell = ws.cell(row=r_idx, column=1).value
            is_total   = str(first_cell or "").strip() == "Complessivo"
            is_alt     = (r_idx % 2 == 0) and not is_total
            _style_row(ws, r_idx, n_cols, is_alt=is_alt, is_total=is_total)

        # Allineamenti e formati numerici
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                col_name = ws.cell(row=1, column=cell.column).value or ""
                if isinstance(cell.value, float):
                    if "Quantità" in col_name or "Prezzo" in col_name:
                        cell.number_format = "#,##0.00000000"
                        cell.alignment = _RIGHT
                    elif "%" in col_name or "IC" in col_name:
                        cell.number_format = "#,##0.00"
                        cell.alignment = _RIGHT
                    else:
                        cell.number_format = "#,##0.00"
                        cell.alignment = _RIGHT
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"
                    cell.alignment = _RIGHT
                elif cell.value is None:
                    cell.alignment = _CENTER
                else:
                    cell.alignment = _LEFT

        # Freeze header
        ws.freeze_panes = "A2"
        _autofit(ws)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
